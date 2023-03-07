##################################################################################
#                       Code by: Diego Garcia Saltori                            #
#                    UTF-8 | Lang: PT-BR | Version: 1.2                          #
##################################################################################
#                       Importar Bibliotecas                                     #
##################################################################################
import sys
import os
import sqlite3
import smtplib
from email.message import EmailMessage
from openpyxl import Workbook
import pandas as pd
from PyQt5 import QtCore
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QStandardItemModel, QStandardItem, QImage, QPalette, QBrush
from PyQt5.QtWidgets import QWidget, QVBoxLayout, QTabWidget, QTableView, QDoubleSpinBox, QLabel, QFrame, QHeaderView
from PyQt5.QtWidgets import QApplication, QFormLayout, QLineEdit, QPushButton, QDateEdit, QSizePolicy, QFileDialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

#Criando banco de dados
conn = sqlite3.connect("Tarefas.db")
cur = conn.cursor()
cur.execute("CREATE TABLE IF NOT EXISTS task (id INTEGER PRIMARY KEY AUTOINCREMENT, task text, date text, value value)")
cur.execute("CREATE TABLE IF NOT EXISTS punish (id INTEGER PRIMARY KEY AUTOINCREMENT, punish text, date text, value value)")
#Classe para criar a aplicação
class App(QWidget):
    #Função para criar a interface do aplicativo
    def __init__(self):
        super().__init__()
        self.title = 'iTask 1.2'
        self.left = 250
        self.top = 100
        self.width = 800
        self.height = 600
        self.initUI()
        self.load_data()
        self.load_data2()        
        self.oImage = QImage("image/fundo.png")
        self.sImage = self.oImage.scaled(self.width, self.height) 
        self.palette = QPalette()
        self.palette.setBrush(10, QBrush(self.oImage))
        #Esse 10 acima é a propriedade Window role, veja o manual da Qt
        self.setPalette(self.palette)
    #Criação das abas da aplicação
    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        #Criação do layout das abas
        layout = QVBoxLayout()
        tabs = QTabWidget() 
        task = QWidget()
        punish = QWidget()
        report = QWidget()
        welcome_tab = QWidget()
        tabs.addTab(welcome_tab, "Bem-Vindo")
        tabs.addTab(task, "Registro de Tarefas")                
        tabs.addTab(punish, "Registro de Multas")
        tabs.addTab(report, "Relatórios")
        layout.addWidget(tabs)  
        #Criando tela de boas vindas
        user = os.getenv("USERNAME")        
        welcome_label = QLabel(user)
        welcome_label.setStyleSheet("font-size: 35px; font-weight: bold;")
        welcome_label.setAlignment(Qt.AlignCenter)
        welcome_label1 = QLabel("Seja bem-vindo!")
        welcome_label1.setStyleSheet("font-size: 30px; font-weight: bold;")
        welcome_label1.setAlignment(Qt.AlignCenter)
        #Criando card de Tarefas
        tarefas_card = QFrame()
        tarefas_card.setFrameShape(QFrame.Box)
        tarefas_card.setLineWidth(1)
        tarefas_layout = QVBoxLayout()
        tarefas_layout.setAlignment(Qt.AlignCenter)
        conn = sqlite3.connect("Tarefas.db")
        st = pd.read_sql_query("SELECT value FROM task", conn)
        soma_tarefas = st['value'].sum()
        tarefas_label = QLabel("Total Tarefas: R$ {:,.2f}".format(float(soma_tarefas)))
        conn.close()
        tarefas_label.setAlignment(Qt.AlignCenter)
        tarefas_card.setFixedSize(450, 80)
        tarefas_card.setStyleSheet("background-color: #20b2aa; color: white; border-radius: 10px; padding: 20px; font-size: 20px;")
        tarefas_layout.addWidget(tarefas_label)
        tarefas_card.setLayout(tarefas_layout)
        #Criando card de multas
        multas_card = QFrame()
        multas_card.setFrameShape(QFrame.Box)
        multas_card.setLineWidth(1)
        multas_layout = QVBoxLayout()
        conn = sqlite3.connect("Tarefas.db")
        sm = pd.read_sql_query("SELECT value FROM punish", conn)
        soma_multas = sm['value'].sum()
        multas_label = QLabel("Total Multas: R$ {:,.2f}".format(float(soma_multas)))
        conn.close()
        multas_label.setAlignment(Qt.AlignCenter)
        multas_card.setFixedSize(450, 80)
        multas_card.setStyleSheet("background-color: #f5001b; color: white; border-radius: 10px; padding: 20px; font-size: 20px;")
        multas_layout.addWidget(multas_label)
        multas_card.setLayout(multas_layout)
        #Criando card de Resultados
        resultado_card = QFrame()
        resultado_card.setFrameShape(QFrame.Box)
        resultado_card.setLineWidth(1)
        resultado_layout = QVBoxLayout()
        conn = sqlite3.connect("Tarefas.db")
        resultado_label = QLabel("Valor a receber: R$ {:,.2f}".format(float(soma_tarefas - soma_multas)))
        conn.close()
        resultado_label.setAlignment(Qt.AlignCenter)
        resultado_card.setFixedSize(450, 80)
        resultado_card.setStyleSheet("background-color: #49b675; color: white; border-radius: 10px; padding: 20px; font-size: 20px;")
        resultado_layout.addWidget(resultado_label)
        resultado_card.setLayout(resultado_layout)
        #Finalizando o Layout da primeira aba de boas vindas
        welcome_layout = QVBoxLayout()
        welcome_layout.setAlignment(Qt.AlignCenter)
        welcome_layout.addWidget(welcome_label)
        welcome_layout.addWidget(welcome_label1)
        welcome_layout.addWidget(tarefas_card)
        welcome_layout.addWidget(multas_card)
        welcome_layout.addWidget(resultado_card)
        send_button = QPushButton("Enviar para meus pais")
        welcome_layout.addWidget(send_button)
        send_button.clicked.connect(self.export_to_excel)
        welcome_tab.setLayout(welcome_layout)
        #Criando o formulario de Tarefas e aba de Tarefas
        form = QFormLayout()
        name = QLineEdit()
        date_field = QDateEdit()
        value_field = QDoubleSpinBox()
        form.addRow("Tarefa:", name)
        date_field.setCalendarPopup(True)
        date_field.setDisplayFormat("dd/MM/yyyy")
        date_field.setDate(QtCore.QDate.currentDate())
        date_field.show()
        form.addRow("Data:", date_field)
        form.addRow("Valor:", value_field)
        #Botão de envio para a tabela
        button = QPushButton("Enviar")       
        button.clicked.connect(lambda: self.update_data(name.text(), date_field.dateTime(), value_field.value()))
        form.addRow(button) 
        #Tabelas de tarefas
        self.data = {'Tarefa': [], 'Data': [], 'Valor': []}
        df = pd.DataFrame(self.data)
        self.table = QTableView()
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  
        self.table.setColumnWidth(0, 100)  # define a largura da coluna 0 como 100 pixels
        self.table.setColumnWidth(1, 150)  # define a largura da coluna 1 como 150 pixels
        self.table.setColumnWidth(2, 200)  # define a largura da coluna 2 como 200 pixels
        self.table.setStyleSheet("""
            QTableView {
                background-color: #fafad2;
                background-attachment: fixed;
                gridline-color: #cbcbcb;
                font-size: 12pt;
                font-family: Arial, sans-serif;
            }
            QTableView::item {
                padding: 5px;
                border: 1px solid #cbcbcb;
                color: black;
            }
            QTableView::item:selected {
                background-color: #cbcbcb;
            }
        """)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.update_dataframe(df)
        form.addRow(self.table)
        #Botão para salvar os dados da tabela no banco de dados
        save_button = QPushButton("Salvar")
        form.addRow(save_button)
        save_button.clicked.connect(self.save_data)
        #Finalizando layout da aba Tarefas
        task.setLayout(form)
        self.show()
        #Criando o formulario de Multas e aba de Multas
        form2 = QFormLayout()
        name2 = QLineEdit()
        date_field2 = QDateEdit()
        value_field2 = QDoubleSpinBox()
        form2.addRow("Multa:", name2)
        date_field2.setCalendarPopup(True)
        date_field2.setDisplayFormat("dd/MM/yyyy")
        date_field2.setDate(QtCore.QDate.currentDate())
        date_field2.show()
        form2.addRow("Data:", date_field2)
        form2.addRow("Valor:", value_field2)
        #Botão de envio para a tabela
        button2 = QPushButton("Enviar")       
        button2.clicked.connect(lambda: self.update_data2(name2.text(), date_field2.dateTime(), value_field2.value()))
        form2.addRow(button2)    
        #Tabelas de Multas
        self.data2 = {'Multa': [], 'Data': [], 'Valor': []}
        df2 = pd.DataFrame(self.data2)
        self.table2 = QTableView()
        self.table2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table2.setColumnWidth(0, 100)  # define a largura da coluna 0 como 100 pixels
        self.table2.setColumnWidth(1, 150)  # define a largura da coluna 1 como 150 pixels
        self.table2.setColumnWidth(2, 200)  # define a largura da coluna 2 como 200 pixels
        self.table2.setStyleSheet("""
            QTableView {
                background-color: #fafad2;
                background-attachment: fixed;
                gridline-color: #cbcbcb;
                font-size: 12pt;
                font-family: Arial, sans-serif;
            }
            QTableView::item {
                padding: 5px;
                border: 1px solid #cbcbcb;
                color: black;
            }
            QTableView::item:selected {
                background-color: #cbcbcb;
            }
        """)
        self.table2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.update_dataframe2(df2)
        #Botão para salvar os dados da tabela no banco de dados
        form2.addRow(self.table2)
        save_button2 = QPushButton("Salvar")
        form2.addRow(save_button2)
        save_button2.clicked.connect(self.save_data2)
        #Finalizando layout da aba Multas        
        punish.setLayout(form2)
        self.show()        
        #Criando a aba de Relatórios e conectando ao banco de dados
        conn = sqlite3.connect("Tarefas.db")
        #Criando variavél para fazer a leitura da coluna de valores para cada tabela do banco Tarefas e Multas
        df = pd.read_sql_query("SELECT value FROM task", conn)
        df2 = pd.read_sql_query("SELECT value FROM punish", conn)
        #Criando variavel para obter a soma total dos valores de cada coluna
        sum_tasks = df['value'].sum()
        sum_punish = df2['value'].sum()
        #Criando gráfico de barras
        fig, ax = plt.subplots()
        plt.bar(x=['Tarefas', 'Multas'], height=[sum_tasks, sum_punish], color=['blue', 'red'])
        #Criando o titulo dos eixos do grafico e nome
        ax.set_title("Meu gráfico de andamento")
        ax.set_xlabel("")
        ax.set_ylabel("Soma dos valores")
        conn.close()
        #Finalizando a inserção do gráfico na aba
        graph = QVBoxLayout(report)
        canvas = FigureCanvas(fig)
        graph.addWidget(canvas)
        report.setLayout(graph)
        self.show()
        self.setLayout(layout)
##################################################################################
#                            Funções para Tarefas                                #
##################################################################################
    #Exportar para Excel, envio de e-mail e Limpar banco de dados
    def export_to_excel(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        fileName, _ = QFileDialog.getSaveFileName(self, "Salvar Arquivo", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if fileName:
            # Conecta ao banco de dados
            conn = sqlite3.connect("Tarefas.db")
            cur = conn.cursor()            
            # Seleciona os dados da tabela 'tasks'
            cur.execute("SELECT id, task, date, value FROM task")
            data1 = cur.fetchall()
            # Cria o DataFrame a partir dos dados
            df1 = pd.DataFrame(data1, columns=["id", "task", "date", "value"])
            # Escreve os dados na worksheet 'Tasks'
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "Tarefas"
            ws1.cell(row=1, column=1).value = "ID"
            ws1.cell(row=1, column=2).value = "Tarefas"
            ws1.cell(row=1, column=3).value = "Data"
            ws1.cell(row=1, column=4).value = "Valor"
            for i in range(len(df1)):
                ws1.cell(row=i+2, column=1).value = df1.loc[i, "id"]
                ws1.cell(row=i+2, column=2).value = df1.loc[i, "task"]
                ws1.cell(row=i+2, column=3).value = df1.loc[i, "date"]
                ws1.cell(row=i+2, column=4).value = df1.loc[i, "value"]
            # Calcula o total da coluna 'value'
            total1 = df1["value"].sum()
            # Seleciona os dados da tabela 'punish'
            cur.execute("SELECT id, punish, date, value FROM punish")
            data2 = cur.fetchall()
            # Cria o DataFrame a partir dos dados
            df2 = pd.DataFrame(data2, columns=["id", "punish", "date", "value"])
            # Escreve os dados na worksheet 'Punish'
            ws2 = wb.create_sheet("Multas")
            ws2.cell(row=1, column=1).value = "ID"
            ws2.cell(row=1, column=2).value = "Multas"
            ws2.cell(row=1, column=3).value = "Data"
            ws2.cell(row=1, column=4).value = "Valor"
            for i in range(len(df2)):
                ws2.cell(row=i+2, column=1).value = df2.loc[i, "id"]
                ws2.cell(row=i+2, column=2).value = df2.loc[i, "punish"]
                ws2.cell(row=i+2, column=3).value = df2.loc[i, "date"]
                ws2.cell(row=i+2, column=4).value = df2.loc[i, "value"]
        # Calcula o total da coluna 'value'
        total2 = df2["value"].sum()
        # Calcula a diferença entre os totais
        diff = total1 - total2
        # Escreve os totais e a diferença na worksheet 'Totais'
        ws3 = wb.create_sheet("Totais")
        ws3.cell(row=1, column=1).value = "Tabela"
        ws3.cell(row=1, column=2).value = "Total"
        ws3.cell(row=2, column=1).value = "Tarefas"
        ws3.cell(row=2, column=2).value = total1
        ws3.cell(row=3, column=1).value = "Multas"
        ws3.cell(row=3, column=2).value = total2
        ws3.cell(row=4, column=1).value = "Mesada"
        ws3.cell(row=4, column=2).value = diff
        # Salva o arquivo
        wb.save(fileName)
        # Configura as informações do servidor SMTP altere os dados para o seu server de email
        smtp_server = "smtp.mail.com"
        smtp_port = 587
        smtp_username = "exemplo@mail.com"
        smtp_password = "exemplo123"
        with smtplib.SMTP(smtp_server, smtp_port) as smtp_conn:
            smtp_conn.ehlo()
            smtp_conn.starttls()
            smtp_conn.login(smtp_username, smtp_password)
            msg = EmailMessage()
            msg["From"] = smtp_username
            msg["To"] = "exemplo@mail.com"
            msg["Subject"] = "Minha mesada foi calculado pelo iTask"
            msg.set_content(f"Mãe e Pai, tudo bem? \nO total da tabela 'Tarefas' é R${total1}.\nO total da tabela 'Multas' é R${total2}. \nE este é o valor que tenho que receber R${diff}. \nUm forte abraço e amo vocês!")
            with open(fileName, "rb") as f:
                file_data = f.read()
                filename = os.path.basename(fileName)
                msg.add_attachment(file_data, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)
            # Envia a mensagem de e-mail
            smtp_conn.send_message(msg)
        # Limpa a tabela 'task' no banco de dados
        cur.execute("DELETE FROM task")
        cur.execute("DELETE FROM punish")
        conn.commit()
        conn.close()
        # Carrega os dados atualizados na interface gráfica
        self.load_data()
        self.load_data2()
    #Enviar tabela para banco de dados
    def update_data(self, name, date_field, value_field):
        self.data['Tarefa'].append(name)
        self.data['Data'].append(date_field.date().toPyDate())
        self.data['Valor'].append(value_field)
        df = pd.DataFrame(self.data)
        self.update_dataframe(df)
    #Fazer update automatico da tabela com o banco ao abrir o aplicativo
    def update_dataframe(self, df):
        model = QStandardItemModel(df.shape[0], df.shape[1])
        model.setHorizontalHeaderLabels(df.columns)
        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                item = QStandardItem(str(df.iloc[i, j]))
                model.setItem(i, j, item)
        self.table.setModel(model)
        self.table.resizeColumnsToContents()
    #Salvar formulário na tabela
    def save_data(self):
        conn = sqlite3.connect("Tarefas.db")
        cur = conn.cursor()
        for i in range(len(self.data['Tarefa'])):
            cur.execute("SELECT * FROM task WHERE task=? AND date=? AND value=?", 
                        (self.data['Tarefa'][i],   
                        self.data['Data'][i], 
                        self.data['Valor'][i]))
            if not cur.fetchone():
                cur.execute("INSERT INTO task (task, date, value) VALUES (?, ?, ?)", 
                            (self.data['Tarefa'][i],  
                            self.data['Data'][i], 
                            self.data['Valor'][i]))
        conn.commit()
        conn.close()
        self.load_data()
    #Carregar automaticamente o banco de dados
    def load_data(self):
        self.data = {'Tarefa': [], 'Data': [], 'Valor': []}
        conn = sqlite3.connect("Tarefas.db")
        cur = conn.cursor()
        cur.execute("SELECT * FROM task")
        rows = cur.fetchall()
        for row in rows:
            self.data['Tarefa'].append(row[1])
            self.data['Data'].append(row[2])
            self.data['Valor'].append(row[3])
        conn.close()
        df = pd.DataFrame(self.data)
        self.update_dataframe(df) 
##################################################################################
#                          Funções para Multas                                   #
##################################################################################
    def update_data2(self, name2, date_field2, value_field2):
        self.data2['Multa'].append(name2)
        self.data2['Data'].append(date_field2.date().toPyDate())
        self.data2['Valor'].append(value_field2)
        df2 = pd.DataFrame(self.data2)
        self.update_dataframe2(df2)
    #Fazer update automatico da tabela com o banco ao abrir o aplicativo 
    def update_dataframe2(self, df2):
        model2 = QStandardItemModel(df2.shape[0], df2.shape[1])
        model2.setHorizontalHeaderLabels(df2.columns)
        for i in range(df2.shape[0]):
            for j in range(df2.shape[1]):
                item2 = QStandardItem(str(df2.iloc[i, j]))
                model2.setItem(i, j, item2)
        self.table2.setModel(model2)
        self.table2.resizeColumnsToContents()
    #Salvar formulário na tabela
    def save_data2(self):
        conn = sqlite3.connect("Tarefas.db")
        cur = conn.cursor()
        for i in range(len(self.data2['Multa'])):
            cur.execute("SELECT * FROM punish WHERE punish=? AND date=? AND value=?", 
                        (self.data2['Multa'][i],  
                        self.data2['Data'][i], 
                        self.data2['Valor'][i]))
            if not cur.fetchone():
                cur.execute("INSERT INTO punish (punish, date, value) VALUES (?, ?, ?)", 
                            (self.data2['Multa'][i],  
                            self.data2['Data'][i], 
                            self.data2['Valor'][i]))
        conn.commit()
        conn.close()
        self.load_data2()
    #Carregar automaticamente o banco de dados
    def load_data2(self):
        self.data2 = {'Multa': [], 'Data': [], 'Valor': []}
        conn = sqlite3.connect("Tarefas.db")
        cur = conn.cursor()
        cur.execute("SELECT * FROM punish")
        rows = cur.fetchall()
        for row in rows:
            self.data2['Multa'].append(row[1])
            self.data2['Data'].append(row[2])
            self.data2['Valor'].append(row[3])
        conn.close()
        df2 = pd.DataFrame(self.data2)
        self.update_dataframe2(df2)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec())